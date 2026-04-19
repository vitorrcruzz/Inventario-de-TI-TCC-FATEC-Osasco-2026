import io
from datetime import datetime
from flask import Blueprint, send_file
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from backend.models import Dispositivo

export_bp = Blueprint("export", __name__)

COLUNAS = [
    "IP", "Hostname", "Tipo", "Status", "Modelo",
    "Sistema Op.", "Processador", "Memória", "Armazenamento",
    "Licença", "Expira em", "Dias Restantes",
    "Localização", "Contato", "Interfaces", "Uptime", "Última Varredura"
]

def _dispositivos_para_df():
    dispositivos = Dispositivo.query.order_by(Dispositivo.ip).all()
    linhas = []
    for d in dispositivos:
        dias = d.dias_para_expirar()
        if dias is None:
            dias_str = "—"
        elif dias < 0:
            dias_str = "EXPIRADA"
        else:
            dias_str = str(dias)
        linhas.append({
            "IP":               d.ip,
            "Hostname":         d.hostname       or "—",
            "Tipo":             d.tipo           or "Desconhecido",
            "Status":           d.status.upper(),
            "Modelo":           d.modelo         or "—",
            "Sistema Op.":      d.sistema_op     or "—",
            "Processador":      d.processador    or "—",
            "Memória":          d.memoria        or "—",
            "Armazenamento":    d.armazenamento  or "—",
            "Licença":          d.licenca_software or "—",
            "Expira em":        d.licenca_expira.strftime("%d/%m/%Y") if d.licenca_expira else "—",
            "Dias Restantes":   dias_str,
            "Localização":      d.localizacao    or "—",
            "Contato":          d.contato        or "—",
            "Interfaces":       d.interfaces     or 0,
            "Uptime":           d.uptime         or "—",
            "Última Varredura": d.ultima_vez.strftime("%d/%m/%Y %H:%M:%S") if d.ultima_vez else "—",
        })
    return pd.DataFrame(linhas, columns=COLUNAS)

@export_bp.route("/export/csv")
def exportar_csv():
    df = _dispositivos_para_df()
    buf = io.StringIO()
    df.to_csv(buf, index=False, sep=";", encoding="utf-8-sig")
    buf.seek(0)
    nome = f"inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return send_file(
        io.BytesIO(buf.getvalue().encode("utf-8-sig")),
        mimetype="text/csv",
        as_attachment=True,
        download_name=nome,
    )

@export_bp.route("/export/excel")
def exportar_excel():
    df = _dispositivos_para_df()
    buf = io.BytesIO()
    num_cols = len(COLUNAS)

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Inventário", startrow=3)
        wb = writer.book
        ws = writer.sheets["Inventário"]

        # ── Título ──
        ws.merge_cells(f"A1:{get_column_letter(num_cols)}2")
        titulo = ws["A1"]
        titulo.value      = "Inventário de Equipamentos de TI — FATEC Osasco"
        titulo.font       = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        titulo.fill       = PatternFill("solid", fgColor="1F4E79")
        titulo.alignment  = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 10

        # ── Cabeçalho ──
        header_fill  = PatternFill("solid", fgColor="2E75B6")
        header_font  = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, col_name in enumerate(COLUNAS, start=1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value     = col_name
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
        ws.row_dimensions[4].height = 30

        # ── Dados ──
        status_cores = {"ONLINE": "E2EFDA", "OFFLINE": "FCE4D6"}
        thin  = Side(style="thin", color="BFBFBF")
        borda = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row_idx in range(5, 5 + len(df)):
            fill_color = "F2F2F2" if (row_idx % 2 == 0) else "FFFFFF"
            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                cell.border    = borda
                if col_idx == 4:  # Status
                    st  = str(cell.value or "")
                    cor = status_cores.get(st, "FFFFFF")
                    cell.fill = PatternFill("solid", fgColor=cor)
                    cell.font = Font(name="Calibri", size=10, bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.fill = PatternFill("solid", fgColor=fill_color)
                    cell.font = Font(name="Calibri", size=10)
            ws.row_dimensions[row_idx].height = 18

        # ── Larguras ──
        larguras = [14, 16, 14, 10, 24, 20, 22, 12, 16, 22, 12, 14, 32, 32, 10, 12, 18]
        for i, w in enumerate(larguras, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── Rodapé ──
        rodape_row = 5 + len(df) + 1
        ws.merge_cells(f"A{rodape_row}:{get_column_letter(num_cols)}{rodape_row}")
        rf = ws.cell(row=rodape_row, column=1)
        rf.value     = f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')} — TCC FATEC Osasco 2026"
        rf.font      = Font(name="Calibri", size=9, italic=True, color="808080")
        rf.alignment = Alignment(horizontal="right")

    buf.seek(0)
    nome = f"inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nome,
    )